import datetime
import openpyxl as oxl
from openpyxl.styles import Alignment, Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side
from openpyxl.utils import range_boundaries
from openpyxl.utils import get_column_letter



thin_border = Side(style='thin', color='000000')
thick_border = Side(style='thick', color='000000')
thin_border_ = Border(
    top=thin_border,
    bottom=thin_border,
    left=thin_border,
    right=thin_border
)
thick_border_ = Border(
    top=thick_border,
    bottom=thick_border,
    left=thick_border,
    right=thick_border
)

def create_head(ws):

    ws['A1'] = "Таблица состояний"
    ws['A2'] = "Последнее обновление:"
    ws['A3'] = "ID"
    ws['B3'] = "Имя проекта"
    ws['C3'] = "Разработчик"
    ws['D3'] = "Статус"
    ws['E3'] = "Последнее обновление"
    ws['F3'] = "БД+Символ"
    ws['G3'] = "Посадочное"
    ws['H3'] = "Переведен в постоянные"

    style_center = Alignment(wrap_text=True, horizontal='center', vertical='center')
    for litera in ["A", "B", "C", "D", "E", "F", "G", "H"]:
        ws[f"{litera}3"].alignment = style_center
        ws[f"{litera}3"].font = Font(bold=True)
        ws[f"{litera}3"].border = thin_border_

    for column_litera in ["A", "C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[column_litera].width = 25
    ws.column_dimensions["B"].width = 35
    ws.row_dimensions[3].height = 32

def set_column_autowidth(ws, columns, reserve=1.2):
    """
    Устанавливает оптимальную ширину столбцов на основе содержимого.
    """
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter  # Получаем букву столбца (A, B, C, ...)
        
        if column in columns:
            # Находим максимальную длину текста в столбце
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        
            # Устанавливаем ширину столбца с небольшим запасом
            adjusted_width = (max_length + 2) * reserve  # Можно изменить коэффициент для более комфортного отображения
            ws.column_dimensions[column].width = adjusted_width


def add_outer_border(ws, cell_range, style='medium', color='000000'):
    """
    Добавляет внешнюю рамку вокруг указанного диапазона ячеек.

    :param ws: рабочий лист (worksheet) openpyxl
    :param cell_range: строка с диапазоном, например 'B2:D4'
    :param style: стиль линии ('thin', 'thick', 'medium' и т.д.)
    :param color: цвет в формате RRGGBB (по умолчанию чёрный)
    """
    # Преобразуем 'B2:D4' → (min_col, min_row, max_col, max_row)
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)

    side = Side(style=style, color=color)

    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            current = cell.border

            # Определяем, какие стороны должны быть обведены
            top    = side if row == min_row else current.top
            bottom = side if row == max_row else current.bottom
            left   = side if col == min_col else current.left
            right  = side if col == max_col else current.right

            cell.border = Border(top=top, bottom=bottom, left=left, right=right)

