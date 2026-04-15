import openpyxl as oxl
from openpyxl.styles import Alignment, Font
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import datetime

from .utils import pasteList, addOuterBorder, setColumnAutowidth
from system import loadData
from system import log

#from functions import get_components, get_dataframe

#from core.services import getDataframe, getComponents

HEAD = [
    ["Таблица состояний",       "",             "",             "",             "",                     "",             "",             ""                          ], 
    ["Последнее обновление:",   "",             "",             "",             "",                     "",             "",             ""                          ],
    ["ID",                      "Имя проекта",  "Разработчик",  "Статус",       "Последнее обновление", "БД+Символ",    "Посадочное",   "Переведен в постоянные"    ]
]

SUB_HEAD = [
    ["Компонент",               "Схемсимвол",   "БД+Символ",    "Посадочное",   "Перевод"   ]
]

HEAD_HEIGHT = 32
NORMAL_HEIGHT = 21
NORMAL_WIDTH = 25
WIDE_WIDTH = 35
TIGHT_WIDTH = 20

STYLE_CENTER = Alignment(wrap_text=True, horizontal='center', vertical='center')
STYLE_LEFT = Alignment(horizontal='left', vertical='center')
STYLE_THIN_BORDER = Side(style='thin', color='000000')
STYLE_THICK_BORDER = Side(style='thick', color='000000')
STYLE_BLUE_TEXT = Font(underline="single", color="0563C1")
THIN_BORDER = Border(
    top = STYLE_THIN_BORDER,
    bottom = STYLE_THIN_BORDER,
    left = STYLE_THIN_BORDER,
    right = STYLE_THIN_BORDER
)
THICK_BORDER = Border(
    top = STYLE_THICK_BORDER,
    bottom = STYLE_THICK_BORDER,
    left = STYLE_THICK_BORDER,
    right = STYLE_THICK_BORDER
)

def createHead(worksheet):
    """
    Creates a header on the page and designs it.

    Args:
        worksheet: Worksheet of workbook.
    """
    pasteList(worksheet, array = HEAD, startPosiiton = "A1")

    for litera in ["A", "B", "C", "D", "E", "F", "G", "H"]:
        worksheet[f"{litera}3"].alignment = STYLE_CENTER
        worksheet[f"{litera}3"].font = Font(bold = True)
        worksheet[f"{litera}3"].border = THIN_BORDER

    for column_litera in ["A", "C", "D", "E", "F", "G", "H"]:
        worksheet.column_dimensions[column_litera].width = NORMAL_WIDTH

    worksheet.column_dimensions["B"].width = WIDE_WIDTH
    worksheet.row_dimensions[3].height = HEAD_HEIGHT





def addSubSheet(workbook):
    pass


def drawXDashboard(config, dataframe, subtables: dict):

    pathDashboardFile = f"{config["GENERAL"]['DEFAULT_PATH']}{config["GENERAL"]['NAME_FILE_DASHBOARD']}"

    yellow = PatternFill(start_color=config['COLORS']['YELLOW'], end_color=config['COLORS']['YELLOW'], fill_type="solid")
    red = PatternFill(start_color=config['COLORS']['RED'], end_color=config['COLORS']['RED'], fill_type="solid")
    blue = PatternFill(start_color=config['COLORS']['BLUE'], end_color=config['COLORS']['BLUE'], fill_type="solid")
    green = PatternFill(start_color=config['COLORS']['GREEN'], end_color=config['COLORS']['GREEN'], fill_type="solid")

    workbook = oxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Сводная таблица"
    createHead(worksheet)

    worksheet['B2'] = datetime.datetime.now()
    worksheet['B2'].alignment = STYLE_LEFT

    main_row = 4
    for dataRow in dataframe.values:
        project = dataRow[1]
        try:
            state = str(dataRow[3])
        except:
            state = "Ошибка"
        match state:
            case "Новый":
                color = yellow
            case "Идет проверка":
                color = red
            case "К переводу":
                color = blue
            case "Готов номинально":
                color = green
            case "Готов":
                color = green
            case "Готов вручную":
                color = green
            case _:
                color = yellow

        for col_idx, value in enumerate(dataRow, start = 1):

            worksheet.cell(row=main_row, column=col_idx, value=value).alignment = STYLE_CENTER
            worksheet.cell(row=main_row, column=col_idx).fill = color
            worksheet.cell(row=main_row, column=col_idx).border = THIN_BORDER
            worksheet.row_dimensions[main_row].height = NORMAL_HEIGHT

            if col_idx == 2:

                new_sheet = workbook.create_sheet(title=value)
                hyperlink_cell1 = new_sheet.cell(row=1, column=1)  
                hyperlink_cell1.value = "К сводной таблице"
                hyperlink_cell1.hyperlink = f"#'{worksheet.title}'!{get_column_letter(col_idx)}{main_row}" 
                hyperlink_cell1.font = Font(underline="single", color="0563C1")  

                new_sheet['A2'] = value
                pasteList(new_sheet, SUB_HEAD, "A3")

                for litera in ["A", "B", "C", "D", "E"]:
                    new_sheet[f"{litera}3"].alignment = STYLE_CENTER
                    new_sheet[f"{litera}3"].font = Font(bold=True)
                    new_sheet[f"{litera}3"].border = THIN_BORDER
                    new_sheet.column_dimensions[litera].width = TIGHT_WIDTH
                new_sheet.row_dimensions[3].height = HEAD_HEIGHT

                sub_row = 4
                for line in subtables[project]:
                    for index, item in enumerate(line, start=1):
                        new_sheet.cell(row=sub_row, column=index, value=item).alignment = STYLE_CENTER
                        new_sheet.cell(row=sub_row, column=index).border = THIN_BORDER
                        new_sheet.row_dimensions[sub_row].height = NORMAL_HEIGHT
                        match item:
                            case "Не проверено":
                                color_ = yellow
                            case "Нет":
                                color_ = red
                            case "Да":
                                color_ = green
                        if item in ["Да", "Нет", "Не проверено"]:
                            new_sheet.cell(row=sub_row, column=index).fill = color_
                    sub_row += 1

                addOuterBorder(new_sheet, f'A3:E{sub_row - 1}')
                setColumnAutowidth(new_sheet, ["B"], reserve = 1.2)

                hyperlink_cell2 = worksheet.cell(row=main_row, column=col_idx)  
                hyperlink_cell2.hyperlink = f"#'{value}'!A1" 
                hyperlink_cell2.font = Font(underline="single", color="0563C1")
                
        main_row += 1

    addOuterBorder(worksheet, f'A3:H{main_row-1}')
    try:
        workbook.save(pathDashboardFile)
    except Exception as e:
        log(f"ERROR: {e}")


# def updateDashboard(config):

#     pathDashboardFile = f"{config["GENERAL"]['DEFAULT_PATH']}{config["GENERAL"]['NAME_FILE_DASHBOARD']}"
#     data = loadData(f"{config["GENERAL"]['DEFAULT_PATH']}{config["GENERAL"]['NAME_FOLDER_DATA']}/{config["GENERAL"]['NAME_FILE_DATA']}")
#     projects = data["PROJECTS"]

#     YELLOW = PatternFill(start_color=config['COLORS']['YELLOW'], end_color=config['COLORS']['YELLOW'], fill_type="solid")
#     RED = PatternFill(start_color=config['COLORS']['RED'], end_color=config['COLORS']['RED'], fill_type="solid")
#     BLUE = PatternFill(start_color=config['COLORS']['BLUE'], end_color=config['COLORS']['BLUE'], fill_type="solid")
#     GREEN = PatternFill(start_color=config['COLORS']['GREEN'], end_color=config['COLORS']['GREEN'], fill_type="solid")

#     dataframe = getDataframe(projects)
#     main_row = 4

#     workbook = oxl.Workbook()
#     worksheet = workbook.active
#     worksheet.title = "Сводная таблица"
#     createHead(worksheet)

#     worksheet['B2'] = datetime.datetime.now()
#     worksheet['B2'].alignment = Alignment(horizontal='left')

#     for dataRow in dataframe.values:
#         project = dataRow[1]
#         match projects[project]["STATE"]:
#             case 0:
#                 color = YELLOW
#             case 4:
#                 color = RED
#             case 5:
#                 color = BLUE
#             case 10:
#                 color = GREEN
#             case 11:
#                 color = GREEN
#             case 12:
#                 color = GREEN
#             case _:
#                 color = YELLOW

#         for col_idx, value in enumerate(dataRow, start = 1):

#             worksheet.cell(row=main_row, column=col_idx, value=value).alignment = STYLE_CENTER
#             worksheet.cell(row=main_row, column=col_idx).fill = color
#             worksheet.cell(row=main_row, column=col_idx).border = THIN_BORDER
#             worksheet.row_dimensions[main_row].height = NORMAL_HEIGHT

#             if col_idx == 2:
#                 new_sheet = workbook.create_sheet(title=value)

#                 hyperlink_cell1 = new_sheet.cell(row=1, column=1)  
#                 hyperlink_cell1.value = "К сводной таблице"
#                 hyperlink_cell1.hyperlink = f"#'{worksheet.title}'!{get_column_letter(col_idx)}{main_row}" 
#                 hyperlink_cell1.font = STYLE_BLUE_TEXT

#                 new_sheet['A2'] = value
#                 pasteList(new_sheet, SUB_HEAD, "A3")

#                 for litera in ["A", "B", "C", "D", "E"]:
#                     new_sheet[f"{litera}3"].alignment = STYLE_CENTER
#                     new_sheet[f"{litera}3"].font = Font(bold=True)
#                     new_sheet[f"{litera}3"].border = THIN_BORDER
#                     new_sheet.column_dimensions[litera].width = TIGHT_WIDTH
#                 new_sheet.row_dimensions[3].height = HEAD_HEIGHT

#                 sub_row = 4
#                 for line in getSubTable(config, projects, project):
#                     for index, item in enumerate(line, start=1):
#                         new_sheet.cell(row=sub_row, column=index, value=item).alignment = STYLE_CENTER
#                         new_sheet.cell(row=sub_row, column=index).border = THIN_BORDER
#                         new_sheet.row_dimensions[sub_row].height = NORMAL_HEIGHT
#                         match item:
#                             case "Не проверено":
#                                 color_ = PatternFill(start_color=config['COLORS']['YELLOW'], end_color=config['COLORS']['YELLOW'], fill_type="solid")
#                             case "Нет":
#                                 color_ = PatternFill(start_color=config['COLORS']['RED'], end_color=config['COLORS']['RED'], fill_type="solid")
#                             case "Да":
#                                 color_ = PatternFill(start_color=config['COLORS']['GREEN'], end_color=config['COLORS']['GREEN'], fill_type="solid")
#                         if item in ["Да", "Нет", "Не проверено"]:
#                             new_sheet.cell(row=sub_row, column=index).fill = color_

#                     sub_row += 1

#                 addOuterBorder(new_sheet, f'A3:E{sub_row-1}')
#                 setColumnAutowidth(new_sheet, ["B"], reserve=1.2)

#                 hyperlink_cell2 = worksheet.cell(row=main_row, column=col_idx)  
#                 hyperlink_cell2.hyperlink = f"#'{value}'!A1" 
#                 hyperlink_cell2.font = STYLE_BLUE_TEXT
                
#         main_row += 1

#     addOuterBorder(worksheet, f'A3:H{main_row-1}')

#     try:
#         workbook.save(pathDashboardFile)
#     except Exception as e:
#         log(f"ERROR: {e}")

from system import options

if __name__ == "__main__":
    
    


    options = options()
    options.print_all_options()
    #updateDashboard(options.config)
    pass