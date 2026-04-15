
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string, range_boundaries



def getRowColumn(position: str):

    try:
        column = column_index_from_string(position[0])
        row = int(position[1:])
    except:
        return 1, 1
    
    return row, column

def pasteList(worksheet, array: list[str], startPosiiton: str = "A1"):
    """
    Inserts a header into the page starting from a specific cell.

    Args:
        worksheet: Worksheet of workbook.
        array: Data to be inserted.
        startPosition: Starting cell (top-left angle of table).
    """
    row, startColumn = getRowColumn(startPosiiton)
    for line in array:
        column = startColumn
        for cell in line:
            worksheet[f"{get_column_letter(column)}{row}"] = cell
            column += 1
        row += 1
    
def setColumnAutowidth(worksheet, columns: list[str], reserve: float = 1.2):
    """
    Sets the optimal column width based on the content.

    Args:
        worksheet: Worksheet.
        columns: List of columns.
        reserve: Reserve coefficient of width.
    """
    for columnCells in worksheet.columns:
        maxLength = 0
        column = columnCells[0].column_letter
        
        if column in columns:
            for cell in columnCells:
                try:
                    if len(str(cell.value)) > maxLength:
                        maxLength = len(str(cell.value))
                except:
                    pass
        
            adjustedWidth = (maxLength + 2) * reserve
            worksheet.column_dimensions[column].width = adjustedWidth

def addOuterBorder(worksheet, cellRange: str, style: str = 'medium', color: str = '000000'):
    """
    Adds an outer border around the range of cells.
    Args:
        worksheet: Worksheet
        cellRange: Range of cells, ex.: 'B2:D4'
        style: Style of line ('thin', 'thick', 'medium' and etc.)
        color: Color in RRGGBB format (default is black)
    """
    min_col, min_row, max_col, max_row = range_boundaries(cellRange)
    side = Side(style=style, color=color)

    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = worksheet.cell(row=row, column=col)
            current = cell.border
            top    = side if row == min_row else current.top
            bottom = side if row == max_row else current.bottom
            left   = side if col == min_col else current.left
            right  = side if col == max_col else current.right
            cell.border = Border(top=top, bottom=bottom, left=left, right=right)
