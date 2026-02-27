import openpyxl as oxl

STATUS_COLUMN = 1
CHECKER_COLUMN = 6
CHECKER_PCB_COLUMN = 11
PN_NUMBER_COLUMN = 2
PROP_COLUMN = 3
PROP_VALUE_COLUMN = 4
PN_CELL = "B1"
MAX_COMPONENTS = 100
MAX_PROPS = 50
MAX_PINS = 2000
COMMENT_COLUMN = 2

import os
import win32com.client
import pythoncom

def calculate_the_formulas(file_path: str, visible: bool = False) -> str:
    """
    Открывает Excel-файл, пересчитывает все формулы и сохраняет изменения.
    
    Функция использует COM-интерфейс Microsoft Excel для принудительного 
    пересчёта формул в файле. 
    """
    # Получаем абсолютный путь к файлу
    abs_path = os.path.abspath(file_path)
    
    # Проверка существования файла
    if not os.path.exists(abs_path):
        raise FileNotFoundError(f"Файл не найден: {abs_path}")
    
    # Проверка расширения
    if not abs_path.lower().endswith(('.xlsx', '.xls', '.xlsm')):
        raise ValueError("Файл должен быть в формате Excel (.xlsx, .xls, .xlsm)")
    
    # Инициализация COM для работы в потоке (если нужно)
    pythoncom.CoInitialize()
    
    excel = None
    try:
        # Запускаем Excel через COM
        excel = win32com.client.Dispatch('Excel.Application')
        excel.Visible = visible
        excel.DisplayAlerts = False  # Отключаем диалоговые окна
        
        # Открываем книгу (Excel автоматически пересчитывает формулы при открытии)
        wb = excel.Workbooks.Open(abs_path)
        
        # Сохраняем и закрываем
        wb.Save()
        wb.Close()
        
        return abs_path
        
    except Exception as e:
        raise RuntimeError(f"Ошибка при пересчёте формул: {str(e)}")
        
    finally:
        # Освобождаем ресурсы
        if excel:
            excel.Quit()
        pythoncom.CoUninitialize()


def find_start_and_end_row(sheet, start=1, end=1000):
    """
    
    Возвращает номера строк начала и конца данных на листе.

    """
    start_row = -1
    end_row = -1

    for row in range(start, end):
        cell_value = sheet.cell(row=row, column=STATUS_COLUMN).value
        if ("Нет" in str(cell_value) and "Да" in str(cell_value) and "Не проверено" in str(cell_value)) or (str(cell_value) in "ДаНетНе проверено") or str(cell_value) == "=A11":
            if start_row == -1:
                start_row = row
        elif start_row != -1:
            end_row = row
            break

    return [start_row, end_row]

def find_row(target, sheet, column, start_row, end_row):

    for row in range(start_row, end_row):
        if sheet.cell(row=row, column=column).value == target:
            return row

    return -1

def combine_checklists(first_path, second_path, output_path):

    warnings = 0

    first_wb = oxl.load_workbook(first_path)
    second_wb = oxl.load_workbook(second_path)

    first_sheet_names = first_wb.sheetnames
    first_main_sheet = first_wb.worksheets[0]
    start_row, end_row = find_start_and_end_row(first_main_sheet)

    second_main_sheet = second_wb.worksheets[0]

    # переносим проверяющих если они не заполнены
    for row in range(start_row, end_row):
        print(1)
        cell_value = first_main_sheet.cell(row=row, column=STATUS_COLUMN).value
        if cell_value != None:
            second_row = find_row(first_main_sheet.cell(row=row, column=PN_NUMBER_COLUMN).value, second_main_sheet, PN_NUMBER_COLUMN, start_row, start_row + MAX_COMPONENTS)
            if second_row != -1 and len(str(second_main_sheet.cell(row=second_row, column=CHECKER_COLUMN).value)) < 2:
                second_main_sheet.cell(row=second_row, column=CHECKER_COLUMN).value = first_main_sheet.cell(row=row, column=CHECKER_COLUMN).value
            else:
                #log("Компонента " + first_main_sheet.cell(row=row, column=2).value + " не обнаружено!", log_object)
                warnings += 1

            if second_row != -1 and len(str(second_main_sheet.cell(row=second_row, column=CHECKER_PCB_COLUMN).value)) < 2:
                second_main_sheet.cell(row=second_row, column=CHECKER_PCB_COLUMN).value = first_main_sheet.cell(row=row, column=CHECKER_PCB_COLUMN).value
            else:
                #log("Компонента " + first_main_sheet.cell(row=row, column=2).value + " не обнаружено!", log_object)
                warnings += 1

        

    for sheet_name in first_sheet_names[1:]:

        print(sheet_name)

        first_sheet = first_wb[sheet_name]
        if sheet_name in second_wb.sheetnames:
            second_sheet = second_wb[sheet_name]
        else:
            #log("Листа " + sheet_name + " не обнаружено!", log_object)
            warnings += 1
            continue

        part_number = first_sheet[PN_CELL].value

        match sheet_name:
            case full_name if full_name == f"{part_number} DB":
                
                f_start_row, f_end_row = find_start_and_end_row(first_sheet)
                s_start_row, s_end_row = find_start_and_end_row(second_sheet)

                for row in range(f_start_row, f_end_row):
                    f_status = first_sheet.cell(row=row, column=STATUS_COLUMN).value
                    if f_status == "Не проверено":
                        continue
                    else:
                        f_prop = first_sheet.cell(row=row, column=PROP_COLUMN).value
                        f_prop_value = first_sheet.cell(row=row, column=PROP_VALUE_COLUMN).value
                        s_row = find_row(f_prop, second_sheet, PROP_COLUMN, s_start_row, s_end_row)
                        if s_row == -1:
                            continue
                        s_status = second_sheet.cell(row=s_row, column=STATUS_COLUMN).value
                        s_prop_value = second_sheet.cell(row=s_row, column=PROP_VALUE_COLUMN).value
                        if s_status == "Не проверено":
                            match f_status:
                                case "Да":
                                    if f_prop_value == s_prop_value:
                                        second_sheet.cell(row=s_row, column=STATUS_COLUMN).value = "Да"
                                    else:
                                        second_sheet.cell(row=s_row, column=STATUS_COLUMN).value = "Нет"
                                        second_sheet.cell(row=s_row, column=COMMENT_COLUMN).value = first_sheet.cell(row=s_row, column=COMMENT_COLUMN).value
                                case "Нет":
                                    if f_prop_value == s_prop_value:
                                        second_sheet.cell(row=s_row, column=STATUS_COLUMN).value = "Нет"
                                        second_sheet.cell(row=s_row, column=COMMENT_COLUMN).value = first_sheet.cell(row=s_row, column=COMMENT_COLUMN).value
                                    else:
                                        second_sheet.cell(row=s_row, column=STATUS_COLUMN).value = "Не проверено"     

            case full_name if full_name == f"{part_number} Sch":
                
                f_start_row, f_end_row = find_start_and_end_row(first_sheet)
                s_start_row, s_end_row = find_start_and_end_row(second_sheet)

                print(f_start_row, f_end_row)
                print(s_start_row, s_end_row)

                for row in range(f_start_row, f_end_row):
                    f_status = first_sheet.cell(row=row, column=STATUS_COLUMN).value
                    if f_status == "Не проверено":
                        continue
                    else:
                        f_prop = first_sheet.cell(row=row, column=PROP_COLUMN).value
                        f_prop_value = first_sheet.cell(row=row, column=PROP_VALUE_COLUMN).value
                        s_row = find_row(f_prop, second_sheet, PROP_COLUMN, s_start_row, s_end_row)
                        if s_row == -1:
                            continue
                        s_status = second_sheet.cell(row=s_row, column=STATUS_COLUMN).value
                        s_prop_value = second_sheet.cell(row=s_row, column=PROP_VALUE_COLUMN).value
                        if s_status == "Не проверено":
                            match f_status:
                                case "Да":
                                    if f_prop_value == s_prop_value:
                                        second_sheet.cell(row=s_row, column=STATUS_COLUMN).value = "Да"
                                    else:
                                        second_sheet.cell(row=s_row, column=STATUS_COLUMN).value = "Нет"
                                        second_sheet.cell(row=s_row, column=COMMENT_COLUMN).value = first_sheet.cell(row=s_row, column=COMMENT_COLUMN).value
                                case "Нет":
                                    if f_prop_value == s_prop_value:
                                        second_sheet.cell(row=s_row, column=STATUS_COLUMN).value = "Нет"
                                        second_sheet.cell(row=s_row, column=COMMENT_COLUMN).value = first_sheet.cell(row=s_row, column=COMMENT_COLUMN).value
                                    else:
                                        second_sheet.cell(row=s_row, column=STATUS_COLUMN).value = "Не проверено"  

                f_start_row, f_end_row = find_start_and_end_row(first_sheet, start=f_end_row)
                s_start_row, s_end_row = find_start_and_end_row(second_sheet, start=s_end_row)

                print(f_start_row, f_end_row)
                print(s_start_row, s_end_row)

                for row in range(f_start_row, f_end_row):
                    f_status = first_sheet.cell(row=row, column=STATUS_COLUMN).value
                    if f_status == "Не проверено":
                        continue
                    else:
                        f_prop = first_sheet.cell(row=row, column=PROP_COLUMN).value
                        f_prop_value = f"{first_sheet.cell(row=row, column=PROP_VALUE_COLUMN).value}{first_sheet.cell(row=row, column=PROP_VALUE_COLUMN+1).value}{first_sheet.cell(row=row, column=PROP_VALUE_COLUMN+2).value}"
                        s_row = find_row(f_prop, second_sheet, PROP_COLUMN, s_start_row, s_end_row)
                        if s_row == -1:
                            continue
                        s_status = second_sheet.cell(row=s_row, column=STATUS_COLUMN).value
                        s_prop_value = f"{second_sheet.cell(row=s_row, column=PROP_VALUE_COLUMN).value}{second_sheet.cell(row=s_row, column=PROP_VALUE_COLUMN+1).value}{second_sheet.cell(row=s_row, column=PROP_VALUE_COLUMN+2).value}"
                        if s_status == "Не проверено":
                            match f_status:
                                case "Да":
                                    if f_prop_value == s_prop_value:
                                        second_sheet.cell(row=s_row, column=STATUS_COLUMN).value = "Да"
                                    else:
                                        second_sheet.cell(row=s_row, column=STATUS_COLUMN).value = "Нет"
                                        second_sheet.cell(row=s_row, column=COMMENT_COLUMN).value = first_sheet.cell(row=s_row, column=COMMENT_COLUMN).value
                                case "Нет":
                                    if f_prop_value == s_prop_value:
                                        second_sheet.cell(row=s_row, column=STATUS_COLUMN).value = "Нет"
                                        second_sheet.cell(row=s_row, column=COMMENT_COLUMN).value = first_sheet.cell(row=s_row, column=COMMENT_COLUMN).value
                                    else:
                                        second_sheet.cell(row=s_row, column=STATUS_COLUMN).value = "Не проверено"  

            case full_name if full_name == f"{part_number} PCB":
                
                f_start_row, f_end_row = find_start_and_end_row(first_sheet)
                s_start_row, s_end_row = find_start_and_end_row(second_sheet)

                for row in range(f_start_row, f_end_row):
                    f_status = first_sheet.cell(row=row, column=STATUS_COLUMN).value
                    if f_status == "Не проверено":
                        continue
                    else:
                        f_prop = first_sheet.cell(row=row, column=PROP_COLUMN).value
                        f_prop_value = first_sheet.cell(row=row, column=PROP_VALUE_COLUMN).value
                        s_row = find_row(f_prop, second_sheet, PROP_COLUMN, s_start_row, s_end_row)
                        if s_row == -1:
                            continue
                        s_status = second_sheet.cell(row=s_row, column=STATUS_COLUMN).value
                        s_prop_value = second_sheet.cell(row=s_row, column=PROP_VALUE_COLUMN).value
                        if s_status == "Не проверено":
                            match f_status:
                                case "Да":
                                    if f_prop_value == s_prop_value:
                                        second_sheet.cell(row=s_row, column=STATUS_COLUMN).value = "Да"
                                    else:
                                        second_sheet.cell(row=s_row, column=STATUS_COLUMN).value = "Нет"
                                        second_sheet.cell(row=s_row, column=COMMENT_COLUMN).value = first_sheet.cell(row=s_row, column=COMMENT_COLUMN).value
                                case "Нет":
                                    if f_prop_value == s_prop_value:
                                        second_sheet.cell(row=s_row, column=STATUS_COLUMN).value = "Нет"
                                        second_sheet.cell(row=s_row, column=COMMENT_COLUMN).value = first_sheet.cell(row=s_row, column=COMMENT_COLUMN).value
                                    else:
                                        second_sheet.cell(row=s_row, column=STATUS_COLUMN).value = "Не проверено"

    second_wb.save(output_path)

    calculate_the_formulas(output_path, False)
    


if __name__ == "__main__":
    
    first_path = "TEST_2026_02_27_11_41_00.xlsx"
    second_path = "TEST_2026_02_27_11_58_36.xlsx"
    

    combine_checklists(first_path, second_path, "OUTPUT_FILE.xlsx")